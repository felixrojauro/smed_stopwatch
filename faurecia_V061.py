#!/usr/bin/python -u
# -*- coding: utf-8 -*-
import sys,  threading, random, Queue, codecs

from PyQt4 import QtGui, QtCore
from datetime import datetime, timedelta
import sys
import time
import threading
# we are not using RPi GPIO in current solution, connecting directly to PLC
#import RPi.GPIO as GPIO
import UI10
import db1_9
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import xlsxwriter

import json
import logging
logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)

import snap7 # for Siemens PLC communication
import S7200 as plc200 # helper for S7 200 alikes
#import S71200 as plc1200 # helper for S7 1200 alikes
import serial # for reading RFID card

import signal
import os
signal.signal(signal.SIGINT, signal.SIG_DFL)

try:
    _fromUtf8 = QtCore.Qunicodeing.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

## RFID Thread - not used since used RFID card reader does not write to serialport, but
##               directly to stdout - just like a normal keyboard

class RfidReadingObject(QtCore.QObject):

    readUser = QtCore.pyqtSignal([str])
    finished = QtCore.pyqtSignal()

    def __init__(self):
        QtCore.QObject.__init__(self)

    def run(self):
        rfid_port = '/dev/pts/28'
        with open('rfid_port') as rfid_port_file:
            rfid_port = rfid_port_file.read().strip()
        rfid_baud = 9600
        rfid = self.openSerialPort(rfid_port, rfid_baud)

        while True:
            reading = rfid.readline().decode()
            if (len(reading) > 0):
                detectedUser = self.detectLoggedUserChange(reading)
                logging.info("Assigned to: %s", detectedUser)
                self.readUser.emit(detectedUser)
            time.sleep(1)

        # emit the finished signal - we're done
        self.finished.emit()


    def detectLoggedUserChange(self, newUserRfid):
        logging.info(newUserRfid)
        with codecs.open('RFid', encoding='utf-8') as employeeFile:
            employeeList = employeeFile.readlines()
            for singleEmployee in employeeList:
                if (singleEmployee.split("::")[0].strip() == newUserRfid.strip()):
                    return singleEmployee.split("::")[1].strip()
        return ""


    def openSerialPort(self, rfid_port, rfid_baud):
        try:
            serialport = serial.Serial(rfid_port, rfid_baud, timeout=0)
        except:
            logging.info("Cannot open %s. Exiting.", rfid_port)
            os.kill(os.getpid(), signal.SIGINT)
        if (serialport.isOpen() == False):
            logging.info("Cannot open %s. Exiting.", rfid_port)
            os.kill(os.getpid(), signal.SIGINT)
        return serialport

## Load image from thread
class LoadImageThread(QtCore.QThread):
  def __init__(self, file, w, h):
    QtCore.QThread.__init__(self)
    self.file = file
    self.w = w
    self.h = h

  def __del__(self):
    self.wait()

  def run(self):
    self.emit(QtCore.SIGNAL('showImage(QString, int, int)'), self.file, self.w, self.h)

class MyWindow(QtGui.QMainWindow):

    loggedInUser = ""
    loggedInUserChanged = QtCore.pyqtSignal('QString', 'bool')

    def initMessageBox(self):
        # Create a custom font (BIG one)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(60)

        self.messageBox = QtGui.QMessageBox()
        self.messageBox.setIcon(QtGui.QMessageBox.Warning)
        self.messageBox.setStandardButtons(QtGui.QMessageBox.Ok)
        self.messageBox.setFont(font)

    #Close App
    def closeEvent(self,event):
        event.accept()
        ui.running= 0


    #Display updated time
    def displayTime(self):
        date = QtCore.QDateTime.currentDateTime().toString(format("dd-MM-yyyy hh:mm:ss"))
        ui.gui.data_hora.setText(date)


    def someFunctionCalledFromAnotherThread(self,image):
        thread = LoadImageThread(file="Imagens/%s.png" % image , w=256, h=256)
        self.connect(thread, QtCore.SIGNAL("showImage(QString, int, int)"), self.showImage)
        thread.start()

    def showImage(self, filename, w, h):
        pixmap = QtGui.QPixmap(filename).scaled(w, h)
        ui.gui.bola.setPixmap(pixmap)




    def rfidTextEditReturnPressed(self):
        detectedUser = self.detectLoggedUserChange(ui.gui.rfid_text_edit.text())
        self.handleLoggedUserChanged(detectedUser)

    def detectLoggedUserChange(self, newUserRfid):
        with codecs.open('RFid', encoding='utf-8') as employeeFile:
            employeeList = employeeFile.readlines()
            for singleEmployee in employeeList:
                if (singleEmployee.split("::")[0].strip() == newUserRfid.trimmed()):
                    return singleEmployee.split("::")[1].strip()
        return ""

    def handleLoggedUserChanged(self, newUser):
        if not newUser:
            logging.info("no user assigned to RFID card: %s", ui.gui.rfid_text_edit.text())
            ui.gui.rfid_text_edit.clear()
            return
        if newUser == self.loggedInUser:
            logging.info("logging out: %s", newUser)
            self.loggedInUser = ""
            ui.gui.data_grid.setVisible(False)
            ui.gui.login_info_box.setVisible(True)
        else:
            logging.info("logging in: %s", newUser)
            self.loggedInUser = newUser
            ui.gui.login_info_box.setVisible(False)
            ui.gui.data_grid.setVisible(True)
        ui.gui.label_logged_in_name.setText(unicode(self.loggedInUser))
        ui.gui.rfid_text_edit.clear()

        self.loggedInUserChanged.emit(self.loggedInUser, len(self.loggedInUser) > 0)


    def showMessageBox(self, text, title):
        self.messageBox.setText(text)
        self.messageBox.setWindowTitle(title)
        return self.messageBox.exec_()




    #init function
    def __init__(self,parent=None):

        super(MyWindow, self).__init__(parent)

        #Start updating timer
        self.timer = QtCore.QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.displayTime)
        self.timer.start()

        #initialize messagebox
        self.initMessageBox()


#Init UI Class
class initUI(QtCore.QObject):


    showMessageBoxSignal = QtCore.pyqtSignal('QString', 'QString')
    messageBoxVisible = False
    input_state = True
    plc = None
#    lastHornStatus = True


    # Inicia a aplicacao global
    # Define queue , timer e inicia loop principal de contagem
    def __init__(self, databaseName, tableName, queryLinesLimit, country, settingsWorksheet):

        super(initUI, self).__init__()
        # Create the queue
        self.queue = Queue.Queue()

        # Configuracao da interface grafica
        # Chama a classe de UI
        self.gui = UI10.Ui_MainWindow()
        self.gui.setupUi(MainWindow)



        export = self.gui.actionExportar_Dados
        export.triggered.connect(lambda: self.export_data())

        languagePT = self.gui.actionPortugues
        languagePT.triggered.connect(lambda: self.defineUI( self.gui, "PT",settingsWorksheet))
        languagePT.triggered.connect(lambda: self.updateMelhores())
        languagePT.triggered.connect(lambda: self.updateUltimos())


        languagePL = self.gui.actionPolaco
        languagePL.triggered.connect(lambda: self.defineUI( self.gui, "PL", settingsWorksheet))
        languagePL.triggered.connect(lambda: self.updateMelhores())
        languagePL.triggered.connect(lambda: self.updateUltimos())

        self.defineUI(self.gui, country, settingsWorksheet)
        self.dbu = dbu





        # Defini timer para correr a leitura de tempo a tempo
        self.timer = QtCore.QTimer()
        QtCore.QObject.connect(self.timer,
                               QtCore.SIGNAL("timeout()"),
                               self.periodico)

        # Inicia o timer com o valor de tempo definido
        self.running = 1

        # Inicia thread de contagem
        self.updateMelhores()
        self.updateUltimos()


        #RFID reader thread - not used since used RFID card reader does not write to serialport, but
        #                     directly to stdout - just like a normal keyboard

#        self.rfidThread = QtCore.QThread()
#        self.rfidWorker = RfidReadingObject()

#        self.rfidWorker.moveToThread(self.rfidThread)
#        self.rfidWorker.finished.connect(self.rfidThread.quit)
#        self.rfidWorker.finished.connect(self.rfidWorker.deleteLater)

#        self.rfidThread.started.connect(self.rfidWorker.run)
#        self.rfidThread.finished.connect(self.rfidWorker.deleteLater)
#        self.rfidThread.finished.connect(self.rfidThread.deleteLater)

#        self.rfidWorker.readUser.connect(MainWindow.handleLoggedUserChanged)

#        self.rfidThread.start()

    def defineUI(self, myUI ,country, settingsWorksheet):
        global dbu
        global tempoObjectivo
        global tempoIntermedio
        global line

        messageBoxText = ""
        messageBoxTitle = ""

        #Depending on language it defines UI

        if country == "PT":
            self.gui.label_melhores.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignCenter)
            MainWindow.setWindowTitle("%s" % (unicode(settingsWorksheet['C2'].value)))
            line = int(settingsWorksheet['C5'].value)
            tempoObjectivo = int(settingsWorksheet['C6'].value)
            tempoIntermedio = int(settingsWorksheet['C7'].value)
            tempoMinimo = int(settingsWorksheet['C8'].value)
            tempoMaximo = int(settingsWorksheet['C9'].value)
            dbu = db1_9.DatabaseUtility(databaseName, tableName, queryLinesLimit, line, tempoObjectivo, tempoIntermedio,
                                        tempoMinimo, tempoMaximo)


            myUI.label_titulo.setText("%s" % (unicode(settingsWorksheet['C3'].value)))
            myUI.linha.setText("% s - %s %s" % (unicode(settingsWorksheet['C4'].value), unicode(settingsWorksheet['H2'].value), unicode(settingsWorksheet['C5'].value)))
            myUI.tempo_objectivo.setText(dbu.obterValorlimite(tempoObjectivo)+ ":000")
            myUI.label_ultimos.setText("%s" % (unicode(settingsWorksheet['H3'].value)))
            myUI.data_hora_ultimos.setText("%s" % (unicode(settingsWorksheet['H4'].value)))
            myUI.data_hora_melhores.setText("%s" % (unicode(settingsWorksheet['H4'].value)))
            myUI.label_melhores.setText("%s" % (unicode(settingsWorksheet['H5'].value)))
            myUI.label_objectivo.setText("%s" % (unicode(settingsWorksheet['H6'].value)))
            myUI.menuFicheiro.setTitle("%s" % (unicode(settingsWorksheet['H7'].value)))
            myUI.actionExportar_Dados.setText("%s" % (unicode(settingsWorksheet['H8'].value)))
            myUI.actionSair.setText("%s" % (unicode(settingsWorksheet['H9'].value)))
            myUI.menuMudar_idioma.setTitle("%s" % (unicode(settingsWorksheet['H12'].value)))
            myUI.actionPortugues.setText("%s" % (unicode(settingsWorksheet['H13'].value)))
            myUI.actionPolaco.setText("%s" % (unicode(settingsWorksheet['H14'].value)))
            myUI.tempo.setText("00:00.000")
            myUI.label_logged_in.setText("%s" % (unicode(settingsWorksheet['H15'].value)))
            self.messageBoxText = "%s" % (unicode(settingsWorksheet['H16'].value))
            self.messageBoxTitle = "%s" % (unicode(settingsWorksheet['H17'].value))
            myUI.label_user_latest.setText("%s" % (unicode(settingsWorksheet['H18'].value)))
            myUI.label_user_best.setText("%s" % (unicode(settingsWorksheet['H18'].value)))
            myUI.login_required_message("%s" % (unicode(settingsWorksheet['H19'].value)))
            self.saveCountry("PT")


        if country == "PL":
            self.gui.label_melhores.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignTop)
            MainWindow.setWindowTitle("%s" % (unicode(settingsWorksheet['D2'].value)))

            line = int(settingsWorksheet['D5'].value)
            tempoObjectivo = int(settingsWorksheet['D6'].value)
            tempoIntermedio = int(settingsWorksheet['D7'].value)
            tempoMinimo = int(settingsWorksheet['D8'].value)
            tempoMaximo = int(settingsWorksheet['D9'].value)
            dbu = db1_9.DatabaseUtility(databaseName, tableName, queryLinesLimit, line, tempoObjectivo, tempoIntermedio,
                                        tempoMinimo, tempoMaximo)


            myUI.label_titulo.setText("%s" % (unicode(settingsWorksheet['D3'].value)))
            myUI.linha.setText(" %s - %s %s" % (unicode(settingsWorksheet['D4'].value), unicode(settingsWorksheet['I2'].value),unicode(settingsWorksheet['D5'].value)))
            myUI.tempo_objectivo.setText(dbu.obterValorlimite(tempoObjectivo) + ":000")
            myUI.label_ultimos.setText("%s" % (unicode(settingsWorksheet['I3'].value)))
            myUI.data_hora_ultimos.setText("%s" % (unicode(settingsWorksheet['I4'].value)))
            myUI.data_hora_melhores.setText("%s" % (unicode(settingsWorksheet['I4'].value)))
            myUI.label_melhores.setText("%s" % (unicode(settingsWorksheet['I5'].value)))
            myUI.label_objectivo.setText("%s" % (unicode(settingsWorksheet['I6'].value)))
            myUI.menuFicheiro.setTitle("%s" % (unicode(settingsWorksheet['I7'].value)))
            myUI.actionExportar_Dados.setText("%s" % (unicode(settingsWorksheet['I8'].value)))
            myUI.actionSair.setText("%s" % (unicode(settingsWorksheet['I9'].value)))
            myUI.menuMudar_idioma.setTitle("%s" % (unicode(settingsWorksheet['I12'].value)))
            myUI.actionPortugues.setText("%s" % (unicode(settingsWorksheet['I13'].value)))
            myUI.actionPolaco.setText("%s" % (unicode(settingsWorksheet['I14'].value)))
            myUI.label_logged_in.setText("%s" % (unicode(settingsWorksheet['I15'].value)))
            myUI.tempo.setText("00:00.000")
            self.messageBoxText = "%s" % (unicode(settingsWorksheet['I16'].value))
            self.messageBoxTitle = "%s" % (unicode(settingsWorksheet['I17'].value))
            myUI.label_user_latest.setText("%s" % (unicode(settingsWorksheet['I18'].value)))
            myUI.label_user_best.setText("%s" % (unicode(settingsWorksheet['I18'].value)))
            myUI.login_required_message.setText("%s" % (unicode(settingsWorksheet['I19'].value)))
            self.saveCountry("PL")



        myUI.label_ultimos.setStyleSheet('color : black')
        myUI.label_melhores.setStyleSheet('color : black')
        myUI.label_user_best.setStyleSheet('color : black')
        myUI.label_user_latest.setStyleSheet('color : black')
        myUI.tabela_ultimos.setColumnWidth(1,200)
        myUI.tabela_melhores.setColumnWidth(1,200)

        myUI.tabela_ultimos.setDisabled(True)
        myUI.tabela_melhores.setDisabled(True)

        myUI.data_grid.setVisible(False)
        myUI.login_info_box.setVisible(True)

    def saveCountry(self, newCountry):

        file = open("country.txt", "w")

        file.write(newCountry)

        file.close()

    def export_data(self):

        #Tabela a exportar + table de colunas
        table = self.dbu.GetTableToExport()
        table2 = self.dbu.Show()

        #data para nome do ficheiro
        date = QtCore.QDateTime.currentDateTime().toString(format("dd_MM_yyyy_hh_mm_ss"))
        workbook = xlsxwriter.Workbook('/home/pi/Desktop/'+unicode(date)+'.xls')

        # nomes das colunas
        sheet = workbook.add_worksheet()
        format1 = workbook.add_format()
        format1.set_num_format(0x0f)
        sheet.set_column('C:C', 13, format1)
        format2 = workbook.add_format()
        format2.set_num_format(0x15)
        sheet.set_column('D:D', 13, format2)
        for r, row in enumerate(table2):
            for c, col in enumerate(row):
                sheet.write(c, r, col)
        # dados
        for r, row in enumerate(table):
            for c, col in enumerate(row):
                sheet.write(r+1, c, col)



    def displayTime(self):
        self.gui.data_hora.setText(QtCore.QDateTime.currentDateTime().toString())     

    def periodico(self):

        ui.entradaProcesso()
        if not self.running:
            app.quit()
            sys.exit(1)

    def entradaProcesso(self):
        global  ultimo
        while self.queue.qsize():
            try:

                msg = self.queue.get(0)
                self.gui.tempo.clear()
                self.gui.tempo.setText(unicode(msg))
            except Queue.Empty:

                pass

    # Conta o tempo
    def contagem(self):
        global inicio
        global temponovo
        global ultimo
        global tempo

        if (aberto == True and inicio == True):

            temponovo = timedelta(seconds=0)
            inicio = False



        ## Calculo diferenca entre leituras separadas por tempo
        tAntes = datetime.now()
        time.sleep(velocidade)
        tDepois = datetime.now()
        ## dferenca = diferenca de tempo de leitura
        diferenca = tDepois - tAntes
        ## tempo0 = valor que encapsula o resto da soma anterior

        tempo0 = temponovo
        temponovo = diferenca + tempo0
        ultimo = temponovo

        #### definicao da cor do semaforo
        segundos= int(ultimo.total_seconds())
        tempo = ultimo
        segundspar = (segundos % 2 == 0)

        if ultimo < timedelta(seconds= tempoObjectivo):
            if segundspar:
                MainWindow.someFunctionCalledFromAnotherThread("green")

            else:
                MainWindow.someFunctionCalledFromAnotherThread("grey")
        elif ultimo < timedelta(seconds=tempoIntermedio):
            if segundspar:
                MainWindow.someFunctionCalledFromAnotherThread("orange")

            else:

                MainWindow.someFunctionCalledFromAnotherThread("grey")

        else:
            if segundspar:
                MainWindow.someFunctionCalledFromAnotherThread("red")
                self.setHorn(True)
            else:

                MainWindow.someFunctionCalledFromAnotherThread("grey")

        s = ultimo.total_seconds()
        st = unicode(ultimo.microseconds)
        st2 = st[:3]
        st4 = time.strftime("%M:%S.", time.gmtime(ultimo.seconds))
        ultimo = st4 + st2
        self.queue.put(ultimo)


    # Verifica o estado do sinal imput e dependendo do mesmo corre a funcao de contagem ou actualiza o ecra

    def connectPlc(self):
        with open('plc_ip') as plc_ip_file:
            plc_ip_read_data = plc_ip_file.read().split(":")
            plc_ip = plc_ip_read_data[0]
            plc_port = int(plc_ip_read_data[1])
            plc_rack = int(plc_ip_read_data[2])
            plc_slot = int(plc_ip_read_data[3])
        logging.info("connecting to PLC: %s, port: %s, rack: %s, slot: %s", plc_ip, plc_port, plc_rack, plc_slot)
        try:
            if (self.plc is not None and self.plc.get_connected() is True):
                self.plc.disconnect()
            self.plc = plc200.S7_200(plc_ip, plc_port, plc_rack, plc_slot, debug=True)
        except:
            logging.info("Cannot connect to plc: %s on port: %s (rack: %s, slot: %s).", plc_ip, plc_port, plc_rack, plc_slot)
            logging.info("Check your plc ip address, port, rack and slot")
            logging.info("Check if you can ping your PLC")
#            os.kill(os.getpid(), signal.SIGINT)
    
    def tempReadInput(self):
        with open('simulate_gpio.json') as gpio_simulator:
            input_var = bool(json.loads(gpio_simulator.read())['bcm'][18])
            return input_var

    def readInput(self):
#        return self.tempReadInput()
        myInput = ""
        local_input_state = self.input_state
        with open('plc_data') as plc_data_file:
            plc_data_list = plc_data_file.readlines()
            myInput = plc_data_list[1]
        try:
            local_input_state = not self.plc.getMem(myInput)
        except:
            logging.info("Failed to read input state. input value left unchanged: %s", self.input_state)
            logging.info("Reconnecting to server now...")
            self.connectPlc()
            return self.input_state
        return local_input_state
    
    def setHorn(self,status):
        myOutput = ""
        with open('plc_data') as plc_data_file:
            plc_data_list = plc_data_file.readlines()
            myOutput = plc_data_list[3]
        try:
#            if (self.lastHornStatus is not status):
            if (self.plc.getMem(myOutput) is not status):
                logging.info("set horn: %s to: %s", myOutput, status)
#                self.lastHornStatus = status
                self.plc.writeMem(myOutput,status)
        except:
            logging.info("Failed to write %s with: %s",myOutput, status)
            logging.info("Reconnecting to server now...")
            self.connectPlc()

    def setRelay(self,status):
        myOutput = ""
        with open('plc_data') as plc_data_file:
            plc_data_list = plc_data_file.readlines()
            myOutput = plc_data_list[5]
        try:
            if (self.plc.getMem(myOutput) is not status):
                logging.info("set relay: %s to: %s", myOutput, status)
                self.plc.writeMem(myOutput,status)
        except:
            logging.info("Failed to write %s with: %s",myOutput, status)
            logging.info("Reconnecting to server now...")
            self.connectPlc()
    
    def messageBoxClosed(self, buttonClicked):
        self.messageBoxVisible = False

    def loggedInUserChanged(self, user, state):
        self.setRelay(state)

    def cronometro(self):

        global plc_ip
        global plc_port
        global plc_rack
        global plc_slot
        global serialport

# just for the record, sample to read/write floating values
#print plc.getMem('freal10')
#plc.writeMem('freal10',3.141592653589)

        self.input_state = self.readInput()
        # Corre se a aplicacao estiver a correr

        while (self.running == 1):

            global aberto
            global inicio
            global ultimo
            global tempo
            global valor
            global test

            while (self.input_state == False and self.running == 1):
                if len(MainWindow.loggedInUser) == 0:
#                    if self.messageBoxVisible == False:
#                        self.showMessageBoxSignal.emit(self.messageBoxText, self.messageBoxTitle)
#                        self.messageBoxVisible = True
                    time.sleep(velocidade)
                else:
                    self.input_state = self.readInput()

                    self.timer.start(80)

                    ## Se troax aberto iniciar tempo
                    if aberto == False:
                        aberto = True

                    # Iniciar contagem de tempo
                    self.contagem()
            while (self.input_state == True and self.running == 1):
                self.input_state = self.readInput()

                self.setHorn(False)
                MainWindow.someFunctionCalledFromAnotherThread("grey")
                ##Se troax fechado
                inicio = True
                aberto = False
                # Actualiza ultimo tempo no interface
                # Limpa tempo actual
                if (ultimo != 0):
                    thread3 = threading.Thread(target=self.insereDB, args=(tempo,))
                    thread3.start()

                ultimo = 0
                if test == True:
                    MainWindow.someFunctionCalledFromAnotherThread("grey")
                    test= False

                time.sleep(velocidade)


    def insereDB(self,  temp):
        global test
        global tempo
        global line
        self.dbu.AddEntryToTable(temp, line, MainWindow.loggedInUser)


        self.updateUltimos()
        self.updateMelhores()
        self.timer.stop()

        test = True





        ######################################### UPDATE DE TABELAS #########################################

    ### Update dos ultimos tempos
    def updateUltimos(self):

        ## Faz query a base de dados dos dados a apresentar na tabela
        tabela = self.dbu.GetTable_last_data("data", "hora", "tempo", "cor", "user")

        ## Obtem um array os valores convertidos para apresentar na tabela
        dados= self.obterArrayDeValoresConvertidos(tabela)

        ## Apresenta os dados na tabela de acordo com as configuracoes de cor, tamabho de letra e centramento
        for valor in range(0, len(dados)):

            itemTempo= self.itemParalinhaDaTabela(0, dados, valor, 19)
            itemTempo.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            self.gui.tabela_ultimos.setItem(valor, 0, itemTempo)

            itemDataHora = self.itemParalinhaDaTabela(1,dados,valor,15)
            itemDataHora.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            self.gui.tabela_ultimos.setItem(valor, 1, itemDataHora)

            itemUser = self.itemParalinhaDaTabela(3, dados, valor, 10)
            itemUser.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            itemUser.setText(self.getSecondPartOfUser((dados[valor])[3]))
#            itemUser.setText(self.getFirstThreeLettersOfUser((dados[valor])[3]))

            self.gui.tabela_ultimos.setItem(valor, 2, itemUser)

    def getFirstThreeLettersOfUser(self, user):
        list = user.split(" ")
        strippedUser = ""
        if (len(list) >= 2):
            strippedUser = list[0][:3] + list[1][:3]
        return strippedUser

    def getSecondPartOfUser(self, user):
        list = user.split(" ")
        strippedUser = ""
        if (len(list) >= 2):
            strippedUser = list[1]
        return strippedUser

    ### Update dos melhores tempos
    def updateMelhores(self):

        ## Faz query a base de dados dos dados a apresentar na tabela
        tabela = self.dbu.GetTable_best_data("data", "hora", "tempo", "cor", "user")
        ## Obtem um array os valores convertidos para apresentar na tabela
        dados = self.obterArrayDeValoresConvertidos(tabela)
        ## Apresenta as linhas na tabela de acordo com as configuracoes de cor, tamabho de letra e centramento
        for valor in range(0, len(dados)):
            itemTempo = self.itemParalinhaDaTabela(0, dados, valor, 19)
            itemTempo.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            self.gui.tabela_melhores.setItem(valor, 0, itemTempo)

            itemDataHora = self.itemParalinhaDaTabela(1, dados, valor, 15)
            itemDataHora.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            self.gui.tabela_melhores.setItem(valor, 1, itemDataHora)

            itemUser = self.itemParalinhaDaTabela(3, dados, valor, 10)
            itemUser.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            itemUser.setText(self.getSecondPartOfUser((dados[valor])[3]))
#            itemUser.setText(self.getFirstThreeLettersOfUser((dados[valor])[3]))

            self.gui.tabela_melhores.setItem(valor, 2, itemUser)


    def obterArrayDeValoresConvertidos(self,tabela):
        array3 = []
        for data, hora, tempo, cor, loggedUser in tabela:
            data = unicode(data.day) + "-" + unicode(data.month) + "-" + unicode(data.year)
            hora = unicode(hora)

            hora = hora[:5]
            if hora.endswith(':'):
                hora = hora[:4]

                hora = "0" + hora
            line = data + "  " + hora

            tempo = unicode(tempo)
            tempo = tempo[:9]
            tempo = tempo.encode('ascii', 'ignore')

            array3.append((unicode(tempo), unicode(line), unicode(cor), unicode(loggedUser)))
        return array3

    def itemParalinhaDaTabela(self, coluna, array, valor, fontSize):

            item2 = QtGui.QTableWidgetItem((array[valor])[coluna])
            tes = ((array[valor])[2])
            item2.setTextColor(my_data[tes])
            font = item2.font()
            font.setPointSize(fontSize)
            item2.setFont(font)
            return item2

# we are not using RPi GPIO in current solution, connecting directly to PLC
# Configuracao GPIO's
#GPIO.setmode(GPIO.BCM)
#GPIO.setup(18, GPIO.IN, pull_up_down=GPIO.PUD_UP)


settingsWorkbook = Workbook()
settingsWorkbook = load_workbook('eSMED_settings.xlsx')
settingsWorkbook.template = True
settingsWorksheet = settingsWorkbook.active

##Variable definition
countryFile = open("country.txt", "rb")
country = str(countryFile.read())
print country
if (country != "PT"):
    if country != "PL":
        country = "PT"

countryFile.close()
countryFile.close()


inicio = True
aberto = False
tempo = 0
test = True
ultimo = 0
velocidade = 0.5
valor = 0
logged_user = ""





#Color definition
green = QtGui.QColor(0,255,0)
red = QtGui.QColor(255,0,0)
yellow = QtGui.QColor(255,128,0)






## Database settings - Name and table
databaseName = 'FaureciaSmed'
tableName = 'Tempos'
queryLinesLimit = 10

my_data = {}
my_data["Verde"] = green
my_data["Vermelho"] = red
my_data["Laranja"] = yellow

## Definition of interface
app = QtGui.QApplication(sys.argv)
## Definition of main window
MainWindow = MyWindow()
text = QtGui.QInputDialog()
ui = initUI(databaseName, tableName, queryLinesLimit, country, settingsWorksheet)
ui.connectPlc()
ui.setHorn(False)
ui.setRelay(False)
ui.showMessageBoxSignal.connect(MainWindow.showMessageBox)

MainWindow.loggedInUserChanged.connect(ui.loggedInUserChanged)
ui.gui.rfid_text_edit.returnPressed.connect(MainWindow.rfidTextEditReturnPressed)

MainWindow.messageBox.buttonClicked.connect(ui.messageBoxClosed)

thread1 = threading.Thread(target=ui.cronometro)
thread1.start()

MainWindow.showMaximized()
sys.exit(app.exec_())


